# author:lihua
# Time:2022/10/29 11:24 AM
import os.path

import openpyxl as openpyxl


class ExcelUtil():
    #获得项目路径
    def get_object_path(self):
        return (os.path.abspath(os.path.dirname(__file__)).split("common")[0])

    def read_excel(self):
        #openpyxl,xlrd
        #加载excel
        wb = openpyxl.load_workbook(self.get_object_path()+"/data/login_data.xlsx")
        #获得sheet对象
        sheet = wb['login']
        #获得excel行数和列数
        print(sheet.max_row,sheet.max_column)
        #循环
        all_list=[]
        for rows in range(2,sheet.max_row+1):
            temp_list=[]
            for column in range(1,sheet.max_column+1):
                temp_list.append(sheet.cell(rows,column).value)
            all_list.append(temp_list)
        return all_list

if __name__ == '__main__':
    ExcelUtil().read_excel()
